import openpyxl, datetime

try:
    wb=openpyxl.load_workbook('saldo.xlsx')
except:
    print('tiedostoa ei voitu avata.. Ei office 2013 xlsx tai ei olemassa tässä folderissa')
    SystemExit

today=datetime.datetime.today()
sh=wb.active
one_day = datetime.timedelta(days=1)
#wb.sheetnames
#sh.title

testi_dict={'100':5000,'175':7,'185':4,'190':7,'198':7,'200':5,'205':5,'210':7,'215':7,'216':7,'217':7,
            '220':5,'222':5,'225':7,'230':21,'233':30,'240':5,'245':7,'253':7,'255':7,'260':5,'265':5,
            '270':5,'560':7,'400':365,'600':1000,'900':1000}

def luku(rivierotin='-----'):
    uusi_taulu2=[] 
    for r in range (2,sh.max_row):
        if sh['F%s' %r].value+(testi_dict.get(sh['A%s' %r].value,14)*one_day)<today:
            for c in range(1,sh.max_column+1):     
                d=sh.cell(row=r,column=c).value
                uusi_taulu2.append(d)
            uusi_taulu2.append(rivierotin)

wb2=openpyxl.Workbook()
wb2.sheetnames
sh2=wb2['Sheet']
sh2.title='Vanhat'

def kirjoitin(rivierotin='-----'):
    r=2
    c=1
    for arvo in uusi_taulu2:
        if arvo != rivierotin:
            sh2.cell(row=r,column=c).value=arvo
            c +=1
        if arvo == rivierotin:
            r +=1
            c = 1
    wb2.save(vanhat.xlsx)
    

if __name__==__main__:
    luku()
    kirjoitin()

